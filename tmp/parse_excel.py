# -*- coding: utf-8 -*-
import openpyxl, io
path = r"C:/Users/User/YandexDisk/Рабочие ФАЙЛЫ/Отдел МАРКЕТИНГА/Проект УСТРОЙКА.РУ Новый/___ У-Стройка_ Структура Каталога (категории и товары) ___.xlsx"
out = io.open(r"C:/Projects/ystroika-website/tmp/excel_head.txt", "w", encoding="utf-8")
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
out.write("SHEETS: " + repr(wb.sheetnames) + "\n")
ws = wb["Лист 1"]
out.write("DIMS: max_row=%s max_col=%s\n" % (ws.max_row, ws.max_column))
for i, r in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    out.write("ROW%d: %s\n" % (i, repr(r)))
out.close()
print("done")
